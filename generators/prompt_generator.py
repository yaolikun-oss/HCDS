#!/usr/bin/env python3
"""HCDS Prompt Generator MVP.

Reads a logical dataset table from .xlsx, .csv, or .tsv and writes deterministic
prompt text. This generator is intentionally conservative: it compiles existing
HCDS fields into observable prompt fragments without adding narrative style.
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

EMPTY_VALUES = {"", "none", "n/a", "na", "null", "unknown", "未知", "无"}

PROHIBITED_OBSERVABLE_TERMS_EN = frozenset({
    "brave",
    "calm",
    "courage",
    "determined",
    "elegant",
    "heroic",
    "intelligence",
    "kindness",
    "loyal",
    "majestic",
    "majesty",
    "natural",
    "noble",
    "powerful",
    "wise",
})

PROHIBITED_OBSERVABLE_TERMS_ZH = frozenset({
    "威严",
    "坚定",
    "悲伤",
    "愤怒",
    "勇敢",
    "霸气",
    "高贵",
    "英雄",
    "自然",
    "优雅",
    "智慧",
    "忠诚",
    "善良",
})

FIELD_ALIASES = {
    "record_id": ("RecordId", "CharacterId", "ID", "Id"),
    "name": ("CanonicalName", "CharacterName", "Name"),
    "dynasty": ("Dynasty",),
    "role": ("HistoricalRole", "OfficialTitle", "Role"),
    "biological_sex": ("BiologicalSex", "Sex"),
    "age": ("EstimatedAge", "AgeGroup", "Age"),
    "face_shape": ("FaceShape",),
    "eye_shape": ("EyeShape",),
    "eye_color": ("EyeColor",),
    "eyebrow_shape": ("EyebrowShape",),
    "nose_shape": ("NoseShape",),
    "mouth_shape": ("MouthShape",),
    "hair_style": ("HairStyle",),
    "hair_length": ("HairLength",),
    "hair_color": ("HairColor",),
    "beard_style": ("BeardStyle",),
    "mustache_style": ("MustacheStyle",),
    "body_build": ("BodyBuild",),
    "height_category": ("HeightCategory",),
    "skin_tone": ("SkinTone",),
    "costume_style": ("CostumeStyle",),
    "headwear": ("Headwear",),
    "upper_garment": ("UpperGarment",),
    "lower_garment": ("LowerGarment",),
    "footwear": ("Footwear",),
    "belt": ("Belt",),
    "accessories": ("Accessories",),
    "armor": ("Armor",),
    "primary_weapon": ("PrimaryWeapon",),
    "secondary_weapon": ("SecondaryWeapon",),
    "body_pose": ("BodyPose",),
    "head_direction": ("HeadDirection",),
    "left_arm": ("LeftArm",),
    "right_arm": ("RightArm",),
    "left_hand": ("LeftHand",),
    "right_hand": ("RightHand",),
    "left_leg": ("LeftLeg",),
    "right_leg": ("RightLeg",),
    "interaction_object": ("InteractionObject",),
    "eye_state": ("EyeState",),
    "eye_direction": ("EyeDirection",),
    "eyebrow_state": ("EyebrowState",),
    "mouth_state": ("MouthState",),
    "facial_tension": ("FacialTension",),
}


@dataclass(frozen=True)
class PromptRecord:
    record_id: str
    prompt: str


def clean(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() in EMPTY_VALUES:
        return ""
    return " ".join(text.split())


def first_value(row: dict[str, object], *names: str) -> str:
    lowered = {key.lower(): key for key in row}
    for name in names:
        key = lowered.get(name.lower())
        if key is not None:
            value = clean(row.get(key))
            if value:
                return value
    return ""


def canonical(row: dict[str, object], alias: str) -> str:
    return first_value(row, *FIELD_ALIASES[alias])



def find_observable_violations(row: dict[str, object]) -> list[str]:
    violations: list[str] = []
    for field_name, raw_value in row.items():
        value = clean(raw_value)
        if not value:
            continue
        lowered = value.lower()
        for term in sorted(PROHIBITED_OBSERVABLE_TERMS_EN):
            if re.search(rf"(?<![A-Za-z]){re.escape(term)}(?![A-Za-z])", lowered):
                violations.append(f"{field_name}={value!r} uses interpretive term {term!r}")
        for term in sorted(PROHIBITED_OBSERVABLE_TERMS_ZH):
            if term in value:
                violations.append(f"{field_name}={value!r} uses interpretive term {term!r}")
    return violations


def validate_observable(row: dict[str, object]) -> None:
    violations = find_observable_violations(row)
    if violations:
        detail = "; ".join(violations)
        raise ValueError(f"Observable Principle violation: {detail}")


def join_values(values: Iterable[str]) -> str:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        value = clean(value)
        if not value:
            continue
        key = value.lower()
        if key in seen:
            continue
        seen.add(key)
        result.append(value)
    return ", ".join(result)


def prefixed(label: str, values: Iterable[str]) -> str:
    joined = join_values(values)
    return f"{label}: {joined}" if joined else ""


def compile_prompt(row: dict[str, object]) -> PromptRecord:
    validate_observable(row)

    record_id = canonical(row, "record_id") or canonical(row, "name") or "record"

    identity = join_values([
        canonical(row, "name"),
        canonical(row, "dynasty"),
        canonical(row, "role"),
    ])
    appearance = prefixed("appearance", [
        canonical(row, "biological_sex"),
        canonical(row, "age"),
        canonical(row, "face_shape"),
        canonical(row, "eye_shape"),
        canonical(row, "eye_color"),
        canonical(row, "eyebrow_shape"),
        canonical(row, "nose_shape"),
        canonical(row, "mouth_shape"),
        canonical(row, "hair_style"),
        canonical(row, "hair_length"),
        canonical(row, "hair_color"),
        canonical(row, "beard_style"),
        canonical(row, "mustache_style"),
        canonical(row, "body_build"),
        canonical(row, "height_category"),
        canonical(row, "skin_tone"),
    ])
    costume = prefixed("costume", [
        canonical(row, "costume_style"),
        canonical(row, "headwear"),
        canonical(row, "upper_garment"),
        canonical(row, "lower_garment"),
        canonical(row, "footwear"),
        canonical(row, "belt"),
        canonical(row, "accessories"),
        canonical(row, "armor"),
        canonical(row, "primary_weapon"),
        canonical(row, "secondary_weapon"),
    ])
    pose = prefixed("pose", [
        canonical(row, "body_pose"),
        canonical(row, "head_direction"),
        canonical(row, "left_arm"),
        canonical(row, "right_arm"),
        canonical(row, "left_hand"),
        canonical(row, "right_hand"),
        canonical(row, "left_leg"),
        canonical(row, "right_leg"),
        canonical(row, "interaction_object"),
    ])
    expression = prefixed("expression", [
        canonical(row, "eye_state"),
        canonical(row, "eye_direction"),
        canonical(row, "eyebrow_state"),
        canonical(row, "mouth_state"),
        canonical(row, "facial_tension"),
    ])

    prompt = join_values([identity, appearance, costume, pose, expression])
    return PromptRecord(record_id=record_id, prompt=prompt)


def read_csv_like(path: Path, delimiter: str) -> list[dict[str, object]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle, delimiter=delimiter)
        if not reader.fieldnames:
            raise ValueError(f"{path} has no header row")
        return [dict(row) for row in reader if any(clean(v) for v in row.values())]


def read_xlsx(path: Path, sheet_name: str | None = None) -> list[dict[str, object]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Reading .xlsx files requires openpyxl") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheets = [workbook[sheet_name]] if sheet_name else workbook.worksheets
    for sheet in sheets:
        rows = list(sheet.iter_rows(values_only=True))
        header_index = None
        headers: list[str] = []
        for index, row in enumerate(rows):
            candidate = [clean(cell) for cell in row]
            if any(candidate):
                header_index = index
                headers = candidate
                break
        if header_index is None:
            continue
        records = []
        for row in rows[header_index + 1 :]:
            values = list(row)
            record = {
                header: values[i] if i < len(values) else ""
                for i, header in enumerate(headers)
                if header
            }
            if any(clean(v) for v in record.values()):
                records.append(record)
        if records:
            return records
    raise ValueError(f"{path} contains no data rows")


def read_dataset(path: Path, sheet_name: str | None = None) -> list[dict[str, object]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return read_csv_like(path, ",")
    if suffix == ".tsv":
        return read_csv_like(path, "\t")
    if suffix == ".xlsx":
        return read_xlsx(path, sheet_name=sheet_name)
    raise ValueError(f"Unsupported input format: {path.suffix}")


def write_prompts(records: list[PromptRecord], output: Path) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    lines: list[str] = []
    for record in records:
        lines.append(f"## {record.record_id}")
        lines.append(record.prompt)
        lines.append("")
    output.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Generate HCDS prompt text from a dataset table.")
    parser.add_argument("--input", required=True, type=Path, help="Input .xlsx, .csv, or .tsv dataset table.")
    parser.add_argument("--output", required=True, type=Path, help="Output prompt .txt path.")
    parser.add_argument("--sheet", help="Optional worksheet name for .xlsx input.")
    return parser


def main(argv: list[str] | None = None) -> int:
    parser = build_arg_parser()
    args = parser.parse_args(argv)
    try:
        rows = read_dataset(args.input, sheet_name=args.sheet)
        prompts = [compile_prompt(row) for row in rows]
        prompts = [record for record in prompts if record.prompt]
        if not prompts:
            raise ValueError("No prompt content could be generated from the dataset")
        write_prompts(prompts, args.output)
    except Exception as exc:
        print(f"error: {exc}", file=sys.stderr)
        return 1
    print(f"generated {len(prompts)} prompt(s): {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
